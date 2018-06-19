import tkinter as tk
from tkinter import ttk
import os
from openpyxl import load_workbook
from datetime import datetime
from git import Repo

dict = {}
base_path = str(os.getcwd()).replace("\\", "/")
excel_path = base_path+'/weeklyReport.xlsx'

win = tk.Tk()
win.title('测试组周报填写')
# 开始时间
stlabel = ttk.Label(win, text='开始时间：')
stlabel.grid(column=0, row=1)
start_time_value = tk.StringVar()
start_time_value_entry = ttk.Entry(win, textvariable=start_time_value)
start_time_value_entry.grid(column=1, row=1)
# 结束时间
ftlabel = ttk.Label(win, text='结束时间：')
ftlabel.grid(column=2, row=1)
finish_time_name = tk.StringVar()
entry_fday_name = ttk.Entry(win, textvariable=finish_time_name)
entry_fday_name.grid(column=3, row=1)
# 选择测试还是配置
ch_var = tk.StringVar()
ch_label = ttk.Label(win, text='配置和测试二选一')
ch_label.grid(column=3, row=7)

person_name = tk.StringVar()
ch_person = ttk.Label(win,text='填报人在最右方选一个')
ch_person.grid(column = 3,row =9)
# 计划人天
planlabel = ttk.Label(win, text='计划人天：')
planlabel.grid(column=0, row=3)
plan_day_name = tk.StringVar()
entry_pday_name = ttk.Entry(win, textvariable=plan_day_name)
entry_pday_name.grid(column=1, row=3)
# 实际人天
finishlabel = ttk.Label(win, text='实际人天：')
finishlabel.grid(column=2, row=3)
finish_day_name = tk.StringVar()
entry_fday_name = ttk.Entry(win, textvariable=finish_day_name)
entry_fday_name.grid(column=3, row=3)
# 本周工作内容
day1 = ttk.Label(win, text='day1：*')
day1.grid(column=0, row=6)
day2 = ttk.Label(win, text='day2：*')
day2.grid(column=0, row=7)
day3 = ttk.Label(win, text='day3：*')
day3.grid(column=0, row=8)
day4 = ttk.Label(win, text='day4：*')
day4.grid(column=0, row=9)
day5 = ttk.Label(win, text='day5：*')
day5.grid(column=0, row=10)
this_week_jobs = ttk.Label(win, text='本周工作内容：')
this_week_jobs.grid(column=2, row=5)
this_week_content = tk.StringVar()
day1_content = tk.Text(win, height=3)
day1_content.grid(column=2, row=6)
day2_content = tk.Text(win, height=3)
day2_content.grid(column=2, row=7)
day3_content = tk.Text(win, height=3)
day3_content.grid(column=2, row=8)
day4_content = tk.Text(win, height=3)
day4_content.grid(column=2, row=9)
day5_content = tk.Text(win, height=3)
day5_content.grid(column=2, row=10)

# 下周工作计划
next_week_jobs = ttk.Label(win, text='下周工作计划：*')
next_week_jobs.grid(column=2, row=11)
next_week_content_edittext = tk.Text(win, width=60, height=5)
next_week_content_edittext.grid(column=2, row=12)

# 下周计划开始时间
next_week_start_time = ttk.Label(win, text='开始时间：(写在下方）')
next_week_start_time.grid(column=1, row=11)
next_week_start_time_text = tk.StringVar()
next_week_start_time_text_entry = ttk.Entry(win, textvariable=next_week_start_time_text)
next_week_start_time_text_entry.grid(column=1, row=12)
# 下周计划结束时间
next_week_finish_time = ttk.Label(win, text='结束时间：(写在下方）')
next_week_finish_time.grid(column=3, row=11)
next_week_finish_time_text = tk.StringVar()
next_week_finish_time_text_entry = ttk.Entry(win, textvariable=next_week_finish_time_text)
next_week_finish_time_text_entry.grid(column=3, row=12)
# 下下周工作计划
next2_week_jobs = ttk.Label(win, text='下下周工作计划：*')
next2_week_jobs.grid(column=2, row=13)
next2_week_content_edittext = tk.Text(win, width=60, height=5)
next2_week_content_edittext.grid(column=2, row=14)

# 下周计划开始时间
next2_week_start_time = ttk.Label(win, text='开始时间：(写在下方）')
next2_week_start_time.grid(column=1, row=13)
next2_week_start_time_text = tk.StringVar()
next2_week_start_time_text_entry = ttk.Entry(win, textvariable=next2_week_start_time_text)
next2_week_start_time_text_entry.grid(column=1, row=14)
# 下周计划结束时间
next2_week_finish_time = ttk.Label(win, text='结束时间：(写在下方）')
next2_week_finish_time.grid(column=3, row=13)
next2_week_finish_time_text = tk.StringVar()
next2_week_finish_time_text_entry = ttk.Entry(win, textvariable=next2_week_finish_time_text)
next2_week_finish_time_text_entry.grid(column=3, row=14)

#获得数据
def doData():
	dict = {'start_time_value':start_time_value.get(),
			'finish_time_name':finish_time_name.get(),
			'test_apply':ch_var.get(),'person_name':person_name.get(),
			'plan_day_name':plan_day_name.get(),'finish_day_name':finish_day_name.get(),
			'day1_content':day1_content.get(index1=1.0,index2=100.0),
			'day2_content':day2_content.get(index1=1.0,index2=100.0),
			'day3_content':day3_content.get(index1=1.0,index2=100.0),
			'day4_content':day4_content.get(index1=1.0,index2=100.0),
			'day5_content':day5_content.get(index1=1.0,index2=100.0),
			'next_week_start_time_text':next_week_start_time_text.get(),
			'next_week_finish_time_text':next_week_finish_time_text.get(),
			'next_week_content_edittext':next_week_content_edittext.get(index1=1.0,index2=100.0),
			'next2_week_start_time_text': next2_week_start_time_text.get(),
			'next2_week_finish_time_text': next2_week_finish_time_text.get(),
			'next2_week_content_edittext': next2_week_content_edittext.get(index1=1.0,index2=100.0)
			}
	print(dict)
	return dict

#点击提交按钮
def clickMe():
	writeExcel(doData())
	#通过git进行成员的数据交互，不方便
	pushExcel()
	action['state'] = 'disabled'
	action['text'] = '提交完关闭即可'
	rp1['state'] = 'disabled'
	rp2['state'] = 'disabled'
	rp3['state'] = 'disabled'
	rp4['state'] = 'disabled'
	rp5['state'] = 'disabled'
	rp6['state'] = 'disabled'

action = tk.Button(win, text="提交周报",fg = 'blue',state = 'disabled',command=lambda:clickMe())
action.grid(column=3, row=8)

#push到github
def	pushExcel():
	repo = Repo(base_path)
	remote = repo.remote()
	repo.git.checkout()
	index = repo.index
	index.add(['weeklyReport.xlsx'])
	index.commit('pushed by GitPython')
	remote.push()

#写到excel里
def writeExcel(data):
	wb = load_workbook(excel_path)
	sheet = wb[data['person_name']]
	n_coordinate = 0
	for row in sheet.iter_rows():  # 找到空白行号
		for cell in row:
			# print(cell.coordinate, cell.value)
			if cell.value is not None:
				n_coordinate = cell.row + 1
	# print(n_coordinate)
	this_time = datetime.now().strftime('%Y/%m/%d')
	for n in range(5):
		sheet.cell(n_coordinate+n,1).value = this_time
		sheet.cell(n_coordinate+n,2).value = data['start_time_value']
		sheet.cell(n_coordinate+n,3).value = data['finish_time_name']
		sheet.cell(n_coordinate+n,4).value = "会展统合服务平台"
		sheet.cell(n_coordinate+n,5).value = data['test_apply']
		sheet.cell(n_coordinate+n,6).value = data['person_name']

		sheet.cell(n_coordinate+n,8).value = data['plan_day_name']
		sheet.cell(n_coordinate+n,9).value = data['finish_day_name']

		if n==0:
			sheet.cell(n_coordinate + n, 7).value = data['day1_content']
			sheet.cell(n_coordinate + n, 10).value = this_time
			sheet.cell(n_coordinate + n, 11).value = data['next_week_start_time_text']
			sheet.cell(n_coordinate + n, 12).value = data['next_week_finish_time_text']
			sheet.cell(n_coordinate + n, 13).value = "会展统合服务平台"
			sheet.cell(n_coordinate + n, 14).value = data['test_apply']
			sheet.cell(n_coordinate + n, 15).value = data['person_name']
			sheet.cell(n_coordinate + n, 16).value = data['next2_week_content_edittext']
			sheet.cell(n_coordinate + n, 17).value = "5"
		elif n==1:
			sheet.cell(n_coordinate + n, 7).value = data['day2_content']
			sheet.cell(n_coordinate + n, 10).value = this_time
			sheet.cell(n_coordinate + n, 11).value = data['next2_week_start_time_text']
			sheet.cell(n_coordinate + n, 12).value = data['next2_week_finish_time_text']
			sheet.cell(n_coordinate + n, 13).value = "会展统合服务平台"
			sheet.cell(n_coordinate + n, 14).value = data['test_apply']
			sheet.cell(n_coordinate + n, 15).value = data['person_name']
			sheet.cell(n_coordinate + n, 16).value = data['next2_week_content_edittext']
			sheet.cell(n_coordinate + n, 17).value = "5"
		elif n==2:
			sheet.cell(n_coordinate + n, 7).value = data['day3_content']
		elif n==3:
			sheet.cell(n_coordinate + n, 7).value = data['day4_content']
		elif n==4:
			sheet.cell(n_coordinate + n, 7).value = data['day5_content']

	wb.save(excel_path)

# 更改配置或者测试
def change_selection():
	ch_label.config(text=ch_var.get())

r1 = tk.Radiobutton(win, text='测试', variable=ch_var, value='测试', command=change_selection)
r1.grid(column=3, row=6)
r2 = tk.Radiobutton(win, text='配置', variable=ch_var, value='配置与维护', command=change_selection)
r2.grid(column=4, row=6)

#选择填报人，必须选择后提交周报的按钮才能使用
def change_person():
	ch_person.config(text=person_name.get())
	action['state'] = 'normal'

rp1 = tk.Radiobutton(win, text='方瑛', variable=person_name, value='方瑛', command=change_person)
rp1.grid(column=5, row=7)
rp2 = tk.Radiobutton(win, text='吴桐', variable=person_name, value='吴桐', command=change_person)
rp2.grid(column=5, row=8)
rp3 = tk.Radiobutton(win, text='彭月', variable=person_name, value='彭月', command=change_person)
rp3.grid(column=5, row=9)
rp4 = tk.Radiobutton(win, text='杨晓龙', variable=person_name, value='杨晓龙', command=change_person)
rp4.grid(column=5, row=10)
rp5 = tk.Radiobutton(win, text='程北洋', variable=person_name, value='程北洋', command=change_person)
rp5.grid(column=5, row=11)
rp6 = tk.Radiobutton(win, text='翟松方', variable=person_name, value='翟松方', command=change_person)
rp6.grid(column=5, row=12)

win.mainloop()
